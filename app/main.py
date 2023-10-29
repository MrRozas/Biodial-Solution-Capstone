import tkinter as tk
from tkinter import filedialog
from PIL import Image, ImageTk
import os  # Agregamos la importación de la biblioteca os
import subprocess  # Agregamos la importación de la biblioteca subprocess
import openpyxl

# Funciones para mostrar mensajes en el área de texto
def mostrar_predicción(producto, valor1, valor2, valor3):
    texto_area.insert("end", f"Los valores de los siguientes 3 meses para el producto {producto} son {valor1}, {valor2} y {valor3}\n")

def mostrar_optimización(categoria, nombre, valor1, valor2):
    if categoria == "producto":
        texto_area.insert("end", f"Los valores de optimización para el producto {nombre} son {valor1} para este mes y {valor2} para el mes siguiente\n")
    if categoria == "insumo":
        texto_area.insert("end", f"Los valores de optimización para el insumo {nombre} son {valor1} para este mes y {valor2} para el mes siguiente\n")
    
# Función para abrir un archivo Excel
def abrir_excel(nombre_archivo):
    if nombre_archivo:
        try:
            # Verifica si el archivo existe en el directorio actual
            if os.path.exists(nombre_archivo):
                # Utiliza la biblioteca subprocess para abrir el archivo Excel con la aplicación predeterminada
                subprocess.Popen(["start", "excel", nombre_archivo], shell=True)
            else:
                texto_area.insert("end", f"El archivo {nombre_archivo} no existe en el directorio actual.\n")
        except Exception as e:
            texto_area.insert("end", f"Error al abrir el archivo Excel: {str(e)}\n")

productos = ['ZZ-A-070', 'ZZ-A-352', 'ZZ-A-353', 'ZZ-A-354', 'ZZ-A-355', 'ZZ-A-356']
insumos = ['Bicarbonato de sodio', 'Cloruro de sodio', 'Cloruro de potasio', 'Cloruro calcio Dihidratado', 'Clorudo de magnesio hexadritado', 'Acido Acético', 'Glucosa Anhidra']

def valores_prediccion(nombre_archivo, nombre_hoja):
    # Abre el archivo Excel
    archivo_excel = openpyxl.load_workbook(nombre_archivo, data_only=True)

    # Selecciona una hoja de cálculo
    hoja = archivo_excel[nombre_hoja]

    # Definir las filas iniciales que deseas recorrer
    filas = [21, 36, 51, 66, 81, 96]
    counter = 0
    # Recorre las filas
    for fila_inicio in filas:
            valores = []
            valores_fila = [round(hoja.cell(row=fila_inicio + offset, column=6).value, 2) for offset in range(3)]
            valores.extend(valores_fila)
            mostrar_predicción(productos[counter], valores[0], valores[1], valores[2])
            counter += 1
    texto_area.insert("end", "\n")

    archivo_excel.close()

def valores_optimizacion(nombre_archivo, nombre_hoja):
    # Abre el archivo Excel
    archivo_excel = openpyxl.load_workbook(nombre_archivo, data_only=True)

    # Selecciona una hoja de cálculo
    hoja = archivo_excel[nombre_hoja]

    counter = 0
    # Recorre las filas
    for fila_inicio in range(15, 26, 2):
            valor_1 = hoja.cell(row=fila_inicio, column=7).value
            valor_2 = hoja.cell(row=fila_inicio+1, column=7).value
            mostrar_optimización("producto", productos[counter], valor_1, valor_2)
            counter += 1
    texto_area.insert("end", "\n")

    counter = 0
    # Recorre las filas
    for fila_inicio in range(28, 41, 2):
            valor_1 = hoja.cell(row=fila_inicio, column=7).value
            valor_2 = hoja.cell(row=fila_inicio+1, column=7).value
            mostrar_optimización("insumo", insumos[counter], valor_1, valor_2)
            counter += 1
    texto_area.insert("end", "\n")

    archivo_excel.close()

def mod_demanda():
    # Crea una nueva ventana emergente
    ventana_valores = tk.Toplevel(ventana)
    ventana_valores.title("Ingresar Valores (0 para mantener el valor actual)")

    # Configura el tamaño de la ventana emergente
    ventana_valores.geometry("500x600")  # Ajusta las dimensiones según tu preferencia

    # Crea un frame para los botones
    frame_botones = tk.Frame(ventana_valores)
    frame_botones.pack(side=tk.BOTTOM, pady=10)  # Ajusta el espaciado vertical según tu preferencia

    # Crea campos de texto para ingresar valores
    etiqueta_valor = tk.Label(ventana_valores, text="Producto a modificar (ZZ-A-070 por ejemplo):")
    etiqueta_valor.pack()
    valor = tk.Entry(ventana_valores)
    valor.pack()

    etiqueta_valor1 = tk.Label(ventana_valores, text="Valor para enero:")
    etiqueta_valor1.pack()
    valor1 = tk.Entry(ventana_valores)
    valor1.pack()

    etiqueta_valor2 = tk.Label(ventana_valores, text="Valor para febrero:")
    etiqueta_valor2.pack()
    valor2 = tk.Entry(ventana_valores)
    valor2.pack()

    etiqueta_valor3 = tk.Label(ventana_valores, text="Valor para marzo:")
    etiqueta_valor3.pack()
    valor3 = tk.Entry(ventana_valores)
    valor3.pack()
    
    etiqueta_valor4 = tk.Label(ventana_valores, text="Valor para abril:")
    etiqueta_valor4.pack()
    valor4 = tk.Entry(ventana_valores)
    valor4.pack()

    etiqueta_valor5 = tk.Label(ventana_valores, text="Valor para mayo:")
    etiqueta_valor5.pack()
    valor5 = tk.Entry(ventana_valores)
    valor5.pack()

    etiqueta_valor6 = tk.Label(ventana_valores, text="Valor para junio:")
    etiqueta_valor6.pack()
    valor6 = tk.Entry(ventana_valores)
    valor6.pack()

    # Crea campos de texto para ingresar valores
    etiqueta_valor7 = tk.Label(ventana_valores, text="Valor para julio:")
    etiqueta_valor7.pack()
    valor7 = tk.Entry(ventana_valores)
    valor7.pack()

    etiqueta_valor8 = tk.Label(ventana_valores, text="Valor para agosto:")
    etiqueta_valor8.pack()
    valor8 = tk.Entry(ventana_valores)
    valor8.pack()

    etiqueta_valor9 = tk.Label(ventana_valores, text="Valor para septiembre:")
    etiqueta_valor9.pack()
    valor9 = tk.Entry(ventana_valores)
    valor9.pack()
    
    etiqueta_valor10 = tk.Label(ventana_valores, text="Valor para octubre:")
    etiqueta_valor10.pack()
    valor10 = tk.Entry(ventana_valores)
    valor10.pack()

    etiqueta_valor11 = tk.Label(ventana_valores, text="Valor para noviembre:")
    etiqueta_valor11.pack()
    valor11 = tk.Entry(ventana_valores)
    valor11.pack()

    etiqueta_valor12 = tk.Label(ventana_valores, text="Valor para diciembre:")
    etiqueta_valor12.pack()
    valor12 = tk.Entry(ventana_valores)
    valor12.pack()

    # Función para guardar los valores ingresados y cerrar la ventana
    def guardar_valores():
        # Abre el archivo Excel
        archivo_excel = openpyxl.load_workbook('Predicción_Dda.xlsx')

        # Selecciona la hoja de cálculo
        hoja = archivo_excel['Demanda Historica']

        counter = 2
        if valor.get() in productos:
            for i in productos:
                if i == valor.get():
                    break
                counter += 1
        else:
            texto_area.insert("end", "Producto no encontrado\n")
            return

        if valor1.get()!="0": (hoja.cell(row=counter, column=12)).value = int(valor1.get())
        if valor2.get()!="0": (hoja.cell(row=counter, column=13)).value = int(valor2.get())
        if valor3.get()!="0": (hoja.cell(row=counter, column=2)).value = int(valor3.get())
        if valor4.get()!="0": (hoja.cell(row=counter, column=3)).value = int(valor4.get())
        if valor5.get()!="0": (hoja.cell(row=counter, column=4)).value = int(valor5.get())
        if valor6.get()!="0": (hoja.cell(row=counter, column=5)).value = int(valor6.get())
        if valor7.get()!="0": (hoja.cell(row=counter, column=6)).value = int(valor7.get())
        if valor8.get()!="0": (hoja.cell(row=counter, column=7)).value = int(valor8.get())
        if valor9.get()!="0": (hoja.cell(row=counter, column=8)).value = int(valor9.get())
        if valor10.get()!="0": (hoja.cell(row=counter, column=9)).value = int(valor10.get())
        if valor11.get()!="0": (hoja.cell(row=counter, column=10)).value = int(valor11.get())
        if valor12.get()!="0": (hoja.cell(row=counter, column=11)).value = int(valor12.get())

        # Guarda los cambios en el archivo Excel
        archivo_excel.save('Predicción_Dda.xlsx')

        # Cierra el archivo Excel
        archivo_excel.close()

        ventana_valores.destroy()  # Cierra la ventana de entrada de valores

    # Botón para guardar los valores
    boton_guardar = tk.Button(frame_botones, text="Guardar", command=guardar_valores, height=2, width=20)
    boton_guardar.pack(side=tk.LEFT, padx=10)  # Ajusta el espaciado horizontal según tu preferencia

    # Botón para salir de la ventana de entrada de valores
    boton_salir = tk.Button(frame_botones, text="Salir", command=ventana_valores.destroy, height=2, width=20)
    boton_salir.pack(side=tk.RIGHT, padx=10)  # Ajusta el espaciado horizontal según tu preferencia

def actualizar_demanda():
    # Crea una nueva ventana emergente
    ventana_valores = tk.Toplevel(ventana)
    ventana_valores.title("Ingresa el mes para actualizar la demanda")

    # Configura el tamaño de la ventana emergente
    ventana_valores.geometry("400x200")  # Ajusta las dimensiones según tu preferencia

    # Crea un frame para los botones
    frame_botones = tk.Frame(ventana_valores)
    frame_botones.pack(side=tk.BOTTOM, pady=10)  # Ajusta el espaciado vertical según tu preferencia

    # Crea campos de texto para ingresar valores
    etiqueta_valor = tk.Label(ventana_valores, text="Mes a pronosticar (octubre, noviembre o diciembre):")
    etiqueta_valor.pack()
    valor = tk.Entry(ventana_valores)
    valor.pack()

    # Función para guardar los valores ingresados y cerrar la ventana
    def guardar_valores():
        # Abre el archivo Excel
        archivo_excel_opt = openpyxl.load_workbook('Modelamiento.xlsx', data_only=True)
        archivo_excel_dda = openpyxl.load_workbook('Predicción_Dda.xlsx', data_only=True)

        # Selecciona la hoja de cálculo
        hoja_opt = archivo_excel_opt['Modelamiento']
        hoja_dda = archivo_excel_dda['Demanda Historica']

        #Fila inicial por producto (marzo) *Llega hasta diciembre
        inicio = [21, 36, 51, 66, 81, 96]
        counter = 4

        if valor.get()=="octubre":
            for i in inicio: 
                (hoja_opt.cell(row=counter, column=6)).value = 0
                (hoja_opt.cell(row=counter, column=7)).value = (hoja_dda.cell(row=i, column=6)).value
                counter += 1

        if valor.get()=="noviembre":
            for i in inicio: 
                (hoja_opt.cell(row=counter, column=6)).value = (hoja_dda.cell(row=i, column=6)).value
                (hoja_opt.cell(row=counter, column=7)).value = (hoja_dda.cell(row=i+1, column=6)).value
                counter += 1

        if valor.get()=="diciembre":
            for i in inicio: 
                (hoja_opt.cell(row=counter, column=6)).value = (hoja_dda.cell(row=i+1, column=6)).value
                (hoja_opt.cell(row=counter, column=7)).value = (hoja_dda.cell(row=i+2, column=6)).value
                counter += 1

        # Guarda los cambios en el archivo Excel
        archivo_excel_opt.save('Modelamiento.xlsx')

        # Cierra el archivo Excel
        archivo_excel_opt.close()
        archivo_excel_dda.close()

        ventana_valores.destroy()  # Cierra la ventana de entrada de valores

    # Botón para guardar los valores
    boton_guardar = tk.Button(frame_botones, text="Guardar", command=guardar_valores, height=2, width=20)
    boton_guardar.pack(side=tk.LEFT, padx=10)  # Ajusta el espaciado horizontal según tu preferencia

    # Botón para salir de la ventana de entrada de valores
    boton_salir = tk.Button(frame_botones, text="Salir", command=ventana_valores.destroy, height=2, width=20)
    boton_salir.pack(side=tk.RIGHT, padx=10)  # Ajusta el espaciado horizontal según tu preferencia

# Función para abrir la ventana de entrada de valores
def mod_inventario():
    # Crea una nueva ventana emergente
    ventana_valores = tk.Toplevel(ventana)
    ventana_valores.title("Ingresar Valores (0 para mantener el valor actual))")
    
    # Configura el tamaño de la ventana emergente
    ventana_valores.geometry("500x300")  # Ajusta las dimensiones según tu preferencia

    # Crea un frame para los botones
    frame_botones = tk.Frame(ventana_valores)
    frame_botones.pack(side=tk.BOTTOM, pady=10)  # Ajusta el espaciado vertical según tu preferencia

    # Crea campos de texto para ingresar valores
    etiqueta_valor1 = tk.Label(ventana_valores, text="Cantidad de ZZ-A-070:")
    etiqueta_valor1.pack()
    valor1 = tk.Entry(ventana_valores)
    valor1.pack()

    etiqueta_valor2 = tk.Label(ventana_valores, text="Cantidad de ZZ-A-352:")
    etiqueta_valor2.pack()
    valor2 = tk.Entry(ventana_valores)
    valor2.pack()

    etiqueta_valor3 = tk.Label(ventana_valores, text="Cantidad de ZZ-A-353:")
    etiqueta_valor3.pack()
    valor3 = tk.Entry(ventana_valores)
    valor3.pack()
    
    etiqueta_valor4 = tk.Label(ventana_valores, text="Cantidad de ZZ-A-354:")
    etiqueta_valor4.pack()
    valor4 = tk.Entry(ventana_valores)
    valor4.pack()

    etiqueta_valor5 = tk.Label(ventana_valores, text="Cantidad de ZZ-A-355:")
    etiqueta_valor5.pack()
    valor5 = tk.Entry(ventana_valores)
    valor5.pack()

    etiqueta_valor6 = tk.Label(ventana_valores, text="Cantidad de ZZ-A-356:")
    etiqueta_valor6.pack()
    valor6 = tk.Entry(ventana_valores)
    valor6.pack()

    # Función para guardar los valores ingresados y cerrar la ventana
    def guardar_valores():
        # Abre el archivo Excel
        archivo_excel = openpyxl.load_workbook('Modelamiento.xlsx')

        # Selecciona la hoja de cálculo
        hoja = archivo_excel['Modelamiento']

        if valor1.get()!="0": (hoja.cell(row=14, column=4)).value = int(valor1.get())
        if valor2.get()!="0": (hoja.cell(row=15, column=4)).value = int(valor2.get())
        if valor3.get()!="0": (hoja.cell(row=16, column=4)).value = int(valor3.get())
        if valor4.get()!="0": (hoja.cell(row=17, column=4)).value = int(valor4.get())
        if valor5.get()!="0": (hoja.cell(row=18, column=4)).value = int(valor5.get())
        if valor6.get()!="0": (hoja.cell(row=19, column=4)).value = int(valor6.get())

        # Guarda los cambios en el archivo Excel
        archivo_excel.save('Modelamiento.xlsx')

        # Cierra el archivo Excel
        archivo_excel.close()

        ventana_valores.destroy()  # Cierra la ventana de entrada de valores

    # Botón para guardar los valores
    boton_guardar = tk.Button(frame_botones, text="Guardar", command=guardar_valores, height=2, width=20)
    boton_guardar.pack(side=tk.LEFT, padx=10)  # Ajusta el espaciado horizontal según tu preferencia

    # Botón para salir de la ventana de entrada de valores
    boton_salir = tk.Button(frame_botones, text="Salir", command=ventana_valores.destroy, height=2, width=20)
    boton_salir.pack(side=tk.RIGHT, padx=10)  # Ajusta el espaciado horizontal según tu preferencia

# Función para cerrar la ventana principal y salir del programa
def salir_programa():
    ventana.quit()
    ventana.destroy()

# Crear la ventana principal
ventana = tk.Tk()

# Crear un frame para la imagen en la parte superior
frame_imagen = tk.Frame(ventana)

# Cargar una imagen inicial en el frame de la imagen
imagen_fondo = Image.open("biodial.jpeg")  # Reemplaza con la ruta de tu imagen de fondo
imagen_fondo = ImageTk.PhotoImage(imagen_fondo)
label_imagen = tk.Label(frame_imagen, image=imagen_fondo)

# Crear un frame para el texto en la parte media
frame_texto = tk.Frame(ventana)

# Crear un área de texto para mostrar los mensajes en el frame de texto
texto_area = tk.Text(frame_texto, wrap=tk.WORD, font=("Helvetica", 12), height=4)  # Cambia el tamaño de fuente aquí

# Crear un frame para los botones en la parte inferior
frame_botones = tk.Frame(ventana)

# Crear los botones con tamaño personalizado en el frame de botones
boton_resultado_dda = tk.Button(frame_botones, text="Ver predicción demanda", command=lambda:valores_prediccion('Predicción_Dda.xlsx', "Demanda Historica"), height=2, width=100)
boton_resultado_opt = tk.Button(frame_botones, text="Ver valores de optimización", command=lambda:valores_optimizacion('Modelamiento.xlsx', "Modelamiento"), height=2, width=100)
boton_excel_modelo = tk.Button(frame_botones, text="Abrir Excel modelo optimización", command=lambda:abrir_excel("Modelamiento.xlsx"), height=2, width=100)
boton_excel_dda = tk.Button(frame_botones, text="Abrir Excel predicción demanda", command=lambda:abrir_excel("Predicción_Dda.xlsx"), height=2, width=100)
boton_mod_inv = tk.Button(frame_botones, text="Modificar inventario", command=mod_inventario, height=2, width=100)
boton_mod_dda = tk.Button(frame_botones, text="Modificar demanda", command=mod_demanda, height=2, width=100)
boton_act = tk.Button(frame_botones, text="Actualizar demanda en el modelo", command=actualizar_demanda, height=2, width=100)
boton_salir = tk.Button(frame_botones, text="Salir", command=salir_programa, height=2, width=100)

def iniciar_programa():
    ventana.title("Biodial Software")

    # Configurar el tamaño de la ventana
    ventana.geometry("1000x1000")  # Cambia las dimensiones según tu preferencia

    frame_imagen.pack(fill=tk.BOTH, expand=True)

    label_imagen.pack(fill=tk.BOTH, expand=True)

    frame_texto.pack(fill=tk.BOTH, expand=True)

    texto_area.pack(fill=tk.BOTH, expand=True)

    frame_botones.pack(fill=tk.BOTH, expand=True)

    # Configurar la geometría de los botones en el frame de botones
    boton_resultado_dda.pack(expand=True)
    boton_resultado_opt.pack(expand=True)
    boton_excel_modelo.pack(expand=True)
    boton_excel_dda.pack(expand=True)
    #boton_abrir.pack(expand=True)
    #boton_abrir_con_ruta.pack(expand=True)
    boton_mod_inv.pack(expand=True)
    boton_mod_dda.pack(expand=True)
    boton_act.pack(expand=True)
    boton_salir.pack(expand=True)

    # Iniciar el bucle de la aplicación
    ventana.mainloop()

if __name__ == "__main__":
    iniciar_programa()
